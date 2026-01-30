# ==============================================================================
# 📊 EXPORTADOR EXCEL - Exportación a Excel con openpyxl
# ==============================================================================
#
# Exporta datos a formato Excel (.xlsx) usando openpyxl.
# Características:
#   - Múltiples hojas
#   - Estilos y colores
#   - Ancho de columnas automático
#   - Cabeceras con formato
#
# ==============================================================================

import io
from typing import Dict, List, Any, Optional
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False
    Workbook = None

from .base_exporter import (
    BaseExporter,
    ExporterFactory,
    FormatoExportacion,
    FiltroExportacion
)


class ExcelExporter(BaseExporter):
    """
    Exportador de datos a formato Excel (.xlsx).
    
    Utiliza openpyxl para generar archivos Excel con:
        - Cabeceras con estilo (color, negrita)
        - Ancho de columnas automático
        - Múltiples hojas
        - Formato de fechas y números
    
    Requisitos:
        pip install openpyxl
    
    Ejemplo:
        exporter = ExcelExporter(nombre_archivo="reporte")
        excel_bytes = exporter.exportar(datos)
        
        # Con múltiples hojas
        exporter.agregar_hoja("Usuarios", datos_usuarios)
        exporter.agregar_hoja("Sesiones", datos_sesiones)
        excel_bytes = exporter.finalizar_libro()
    """
    
    # Colores predefinidos (tema marítimo de MentxuApp)
    COLORES = {
        'cabecera': '1E3A5F',      # Azul marino oscuro
        'cabecera_texto': 'FFFFFF', # Blanco
        'fila_par': 'E8F4FD',       # Azul muy claro
        'fila_impar': 'FFFFFF',     # Blanco
        'exito': '28A745',          # Verde
        'error': 'DC3545',          # Rojo
        'advertencia': 'FFC107'     # Amarillo
    }
    
    def __init__(
        self,
        nombre_archivo: str = "exportacion",
        filtros: FiltroExportacion = None,
        nombre_hoja: str = "Datos",
        aplicar_estilos: bool = True
    ):
        """
        Inicializa el exportador Excel.
        
        Args:
            nombre_archivo: Nombre base del archivo
            filtros: Filtros de exportación
            nombre_hoja: Nombre de la hoja principal
            aplicar_estilos: Si aplicar formato a las celdas
        """
        if not OPENPYXL_DISPONIBLE:
            raise ImportError(
                "openpyxl no está instalado. "
                "Ejecuta: pip install openpyxl"
            )
        
        super().__init__(nombre_archivo, filtros)
        self.nombre_hoja = nombre_hoja
        self.aplicar_estilos = aplicar_estilos
        self._workbook: Optional[Workbook] = None
        self._worksheet = None
        self._fila_actual = 1
        self._campos = []
        self._anchos_columnas = {}
    
    # --- Implementación de métodos abstractos ---
    
    def get_formato(self) -> FormatoExportacion:
        return FormatoExportacion.EXCEL
    
    def get_extension(self) -> str:
        return "xlsx"
    
    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    def _inicializar_buffer(self) -> None:
        """Crea el libro de Excel."""
        self._workbook = Workbook()
        self._worksheet = self._workbook.active
        self._worksheet.title = self.nombre_hoja
        self._fila_actual = 1
        self._anchos_columnas = {}
    
    def _escribir_encabezados(self, campos: List[str]) -> None:
        """Escribe y formatea los encabezados."""
        self._campos = campos
        
        for col, campo in enumerate(campos, 1):
            celda = self._worksheet.cell(row=1, column=col)
            nombre_legible = self._formatear_encabezado(campo)
            celda.value = nombre_legible
            
            # Actualizar ancho
            self._actualizar_ancho(col, nombre_legible)
            
            # Aplicar estilo
            if self.aplicar_estilos:
                self._aplicar_estilo_cabecera(celda)
        
        self._fila_actual = 2
    
    def _escribir_fila(self, datos: Dict[str, Any]) -> None:
        """Escribe una fila de datos."""
        for col, campo in enumerate(self._campos, 1):
            valor = datos.get(campo, "")
            valor_formateado = self._formatear_valor(valor)
            
            celda = self._worksheet.cell(row=self._fila_actual, column=col)
            celda.value = valor_formateado
            
            # Actualizar ancho
            self._actualizar_ancho(col, str(valor_formateado))
            
            # Aplicar estilo de fila
            if self.aplicar_estilos:
                self._aplicar_estilo_fila(celda, self._fila_actual)
        
        self._fila_actual += 1
    
    def _finalizar(self) -> bytes:
        """Ajusta anchos y devuelve bytes del Excel."""
        # Ajustar anchos de columnas
        for col, ancho in self._anchos_columnas.items():
            letra = get_column_letter(col)
            # Añadir margen
            self._worksheet.column_dimensions[letra].width = min(ancho + 2, 50)
        
        # Congelar primera fila
        self._worksheet.freeze_panes = 'A2'
        
        # Guardar en buffer
        buffer = io.BytesIO()
        self._workbook.save(buffer)
        buffer.seek(0)
        
        return buffer.read()
    
    # --- Métodos auxiliares ---
    
    def _formatear_encabezado(self, campo: str) -> str:
        """Convierte nombre de campo a encabezado legible."""
        palabras = campo.replace('_', ' ').split()
        return ' '.join(p.title() for p in palabras)
    
    def _formatear_valor(self, valor: Any) -> Any:
        """Formatea valor para Excel."""
        if valor is None:
            return ""
        
        if isinstance(valor, bool):
            return "Sí" if valor else "No"
        
        if isinstance(valor, list):
            return ", ".join(str(v) for v in valor)
        
        if isinstance(valor, dict):
            import json
            return json.dumps(valor, ensure_ascii=False)
        
        # Intentar convertir fechas string a datetime
        if isinstance(valor, str) and 'T' in valor:
            try:
                return datetime.fromisoformat(valor.replace('Z', '+00:00'))
            except:
                pass
        
        return valor
    
    def _actualizar_ancho(self, col: int, texto: str) -> None:
        """Actualiza el ancho máximo de una columna."""
        ancho_actual = self._anchos_columnas.get(col, 0)
        nuevo_ancho = len(str(texto))
        self._anchos_columnas[col] = max(ancho_actual, nuevo_ancho)
    
    def _aplicar_estilo_cabecera(self, celda) -> None:
        """Aplica estilo a celda de cabecera."""
        celda.font = Font(
            bold=True,
            color=self.COLORES['cabecera_texto']
        )
        celda.fill = PatternFill(
            start_color=self.COLORES['cabecera'],
            end_color=self.COLORES['cabecera'],
            fill_type='solid'
        )
        celda.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        celda.border = Border(
            bottom=Side(style='thin', color='000000')
        )
    
    def _aplicar_estilo_fila(self, celda, fila: int) -> None:
        """Aplica estilo alternado a filas."""
        color = (
            self.COLORES['fila_par'] 
            if fila % 2 == 0 
            else self.COLORES['fila_impar']
        )
        
        celda.fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type='solid'
        )
        celda.alignment = Alignment(vertical='center')
    
    # --- Métodos para múltiples hojas ---
    
    def agregar_hoja(
        self,
        nombre: str,
        datos: List[Dict[str, Any]],
        campos: List[str] = None
    ) -> None:
        """
        Agrega una nueva hoja con datos.
        
        Args:
            nombre: Nombre de la hoja
            datos: Lista de diccionarios
            campos: Campos a incluir
        """
        if self._workbook is None:
            self._workbook = Workbook()
            # Eliminar hoja por defecto si es la primera
            if len(self._workbook.worksheets) == 1:
                self._workbook.remove(self._workbook.active)
        
        # Crear nueva hoja
        self._worksheet = self._workbook.create_sheet(title=nombre)
        self._fila_actual = 1
        self._anchos_columnas = {}
        
        # Determinar campos
        if not campos and datos:
            campos = list(datos[0].keys())
        elif not campos:
            campos = []
        
        # Escribir encabezados
        self._escribir_encabezados(campos)
        
        # Escribir datos
        for fila in datos:
            fila_filtrada = {k: fila.get(k) for k in campos}
            self._escribir_fila(fila_filtrada)
        
        # Ajustar anchos
        for col, ancho in self._anchos_columnas.items():
            letra = get_column_letter(col)
            self._worksheet.column_dimensions[letra].width = min(ancho + 2, 50)
        
        # Congelar primera fila
        self._worksheet.freeze_panes = 'A2'
    
    def agregar_hoja_resumen(self, estadisticas: Dict[str, Any]) -> None:
        """
        Agrega una hoja de resumen con estadísticas.
        
        Args:
            estadisticas: Diccionario con estadísticas
        """
        if self._workbook is None:
            self._workbook = Workbook()
        
        ws = self._workbook.create_sheet(title="Resumen", index=0)
        
        # Título
        ws.merge_cells('A1:B1')
        ws['A1'] = "📊 Resumen de Exportación"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Metadatos
        fila = 3
        for clave, valor in self.metadatos.items():
            ws.cell(row=fila, column=1, value=clave.replace('_', ' ').title())
            ws.cell(row=fila, column=2, value=str(valor))
            fila += 1
        
        fila += 1
        ws.cell(row=fila, column=1, value="Estadísticas").font = Font(bold=True)
        fila += 1
        
        # Estadísticas
        for clave, valor in estadisticas.items():
            ws.cell(row=fila, column=1, value=clave.replace('_', ' ').title())
            ws.cell(row=fila, column=2, value=str(valor))
            fila += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    
    def finalizar_libro(self) -> bytes:
        """
        Finaliza el libro con múltiples hojas.
        
        Returns:
            Bytes del archivo Excel
        """
        if self._workbook is None:
            raise ValueError("No hay libro creado")
        
        buffer = io.BytesIO()
        self._workbook.save(buffer)
        buffer.seek(0)
        
        return buffer.read()


# Registrar el exportador en la Factory (solo si openpyxl está disponible)
if OPENPYXL_DISPONIBLE:
    ExporterFactory.registrar(FormatoExportacion.EXCEL, ExcelExporter)
